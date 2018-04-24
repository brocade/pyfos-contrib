#!/usr/bin/env python3
# Copyright 2018 Brocade Communications Systems LLC.  All rights reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may also obtain a copy of the License at
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

# This application will generate a configuration spreadsheet
# from the current state of the fabric.  The IP address of the
# switch to be accessed is found in the environment variable SWITCHIP
# The username and password are hardcoded below since this was demo code.

import pyfos
import pyfos.pyfos_auth as pyfos_auth
import pyfos.pyfos_util as pyfos_util
import pyfos.pyfos_zone as pyfos_zone
import sys
from pyfos.utils import brcd_util
import json
import pprint
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles.colors import RED
import time
import os

isHttps = "0"
VERBOSE = False
ONLINE = True

class json_encoder(json.JSONEncoder):
    def default(self, obj):
        if hasattr(obj, 'reprJSON'):
            return obj.reprJSON()
        else:
            return json.JSONEncoder.default(self, obj)


def get_session(ipaddr, username, password, vfid):
	session = pyfos_auth.login(username, password, ipaddr, isHttps)
	if pyfos_auth.is_failed_login(session):
		print("login failed because",
			  session.get(pyfos_auth.CREDENTIAL_KEY)
			  [pyfos_auth.LOGIN_ERROR_KEY])
		sys.exit()

	brcd_util.exit_register(session)

	pyfos_auth.vfid_set(session, vfid)
	return session


def pull_data(ipaddr, username, password):
	if (ONLINE):
		session = get_session(ipaddr, username, password, int(os.environ.get("VFID")))
		defined_zone = pyfos_zone.defined_configuration.get(session)
		effective_zone = pyfos_zone.effective_configuration.get(session)
		pyfos_auth.logout(session)
	else:
		file = open("def.json","r")
		defined_zone = json.load(file)
		file.close
		file = open("eff.json","r")
		effective_zone = json.load(file)
		file.close
	return (defined_zone, effective_zone)

def gen_spreadsheet(defined_zone, effective_zone):


	print(json.dumps(effective_zone, cls=json_encoder))
	defined_zone = json.loads(json.dumps(defined_zone, cls=json_encoder, sort_keys=True))
	effective_zone = json.loads(json.dumps(effective_zone, cls=json_encoder, sort_keys=True))

	#Set up worksheets
	wb = Workbook()
	wb.active.title = "Effective Zones"
	wb.create_sheet(title="Defined Configurations")
	wb.create_sheet(title="Defined Zones")
	wb.create_sheet(title="Defined Aliases")	




	sheet = wb.get_sheet_by_name("Defined Configurations")
	for i in range(1,50):
		sheet.column_dimensions[get_column_letter(i)].width = 20

	sheet['A1'].font = Font(bold=True)
	sheet['A1'] = "Configuration Name"
	sheet['B1'].font = Font(bold=True)
	sheet['B1'] = "Zone Names"
	theRow = 2
	if defined_zone['defined-configuration']:
		for zoneConfig in defined_zone['defined-configuration']['cfg']:
			sheet.cell(row = theRow, column=1).value = (zoneConfig["cfg-name"])
			for i in range(len(zoneConfig["member-zone"]["zone-name"])):
				sheet.cell(row = theRow, column = (i+2)).value = (zoneConfig["member-zone"]["zone-name"][i])
			theRow = theRow + 1

	sheet = wb.get_sheet_by_name("Defined Zones")
	sheet['A1'].font = Font(bold=True)
	sheet['A1'] = "Zone Name"
	sheet['A2'].font = Font(bold=True)
	sheet['A2'] = "Members"
	for i in range(1,50):
		sheet.column_dimensions[get_column_letter(i)].width = 23
	theColumn = 2
	if defined_zone['defined-configuration']:
		for zoneIndex in range(len(defined_zone['defined-configuration']['zone'])):
			theZone = defined_zone['defined-configuration']['zone'][zoneIndex]
			sheet.cell(row=1, column=theColumn).value = theZone["zone-name"]
			for i in range(len(theZone["member-entry"]["entry-name"])):
				sheet.cell(row=(i+2), column = theColumn).value = theZone["member-entry"]["entry-name"][i]
			theColumn = theColumn + 1
	sheet = wb.get_sheet_by_name("Defined Aliases")
	sheet['A1'].font = Font(bold=True)
	sheet['A1'] = "Alias Name"
	sheet['B1'].font = Font(bold=True)
	sheet['B1'] = "Member Entry"
	for i in range(1,50):
		sheet.column_dimensions[get_column_letter(i)].width = 23
	theRow = 2
	if defined_zone['defined-configuration']:
		for aliasIndex in range(len(defined_zone['defined-configuration']['alias'])):
			theAlias = defined_zone['defined-configuration']['alias'][aliasIndex]
			sheet.cell(row=theRow, column=1).value = theAlias["alias-name"]
			for i in range(len(theAlias["member-entry"]["alias-entry-name"])):
				sheet.cell(row=theRow, column = (i + 2)).value = theAlias["member-entry"]["alias-entry-name"][i]
			theRow = theRow + 1

	sheet = wb.get_sheet_by_name("Effective Zones")
	sheet['A1'].font = Font(bold=True)
	sheet['A1'] = "Effective Configuration"
	sheet['B1'].font = Font(color=RED)
	if "cfg-name" in effective_zone['effective-configuration'].keys():
		sheet['B1'] = effective_zone["effective-configuration"]["cfg-name"]

	sheet['A3'].font = Font(bold=True)
	sheet['A4'].font = Font(bold=True)
	sheet['A3'] = "Zone Name"
	sheet['A4'] = "Members"
	for i in range(1,50):
		sheet.column_dimensions[get_column_letter(i)].width = 23
	theColumn = 2
	if "enabled-zone" in effective_zone['effective-configuration'].keys():
		for zoneIndex in range(len(effective_zone['effective-configuration']['enabled-zone'])):
			theZone = effective_zone['effective-configuration']['enabled-zone'][zoneIndex]
			sheet.cell(row=3, column=theColumn).value = theZone["zone-name"]
			for i in range(len(theZone["member-entry"]["entry-name"])):
				sheet.cell(row=(i+4), column = theColumn).value = theZone["member-entry"]["entry-name"][i]
			theColumn = theColumn + 1

	wb.save("results.xlsx")

def main():

	# Gather switch/fabric information
	defined_zone, effective_zone = pull_data(os.environ.get('SWITCHIP'), "admin", "password")

	# Generate the spreadsheet
	gen_spreadsheet(defined_zone, effective_zone)

	return time.strftime("%Y-%m-%d-%H-%M-%S")



if (__name__ == "__main__"):
	main()
